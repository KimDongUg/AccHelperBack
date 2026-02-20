"""Super admin endpoints for tenant and quota management."""

import io
import logging

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.database import get_db
from app.dependencies import require_super_admin
from app.models.company import Company
from app.models.qa_knowledge import QaKnowledge
from app.models.tenant_quota import TenantQuota
from app.services.embedding_service import bulk_rebuild_embeddings, upsert_qa_embedding

logger = logging.getLogger("acchelper")

router = APIRouter(prefix="/super", tags=["super-admin"])


class TenantUpdate(BaseModel):
    status: str | None = None  # active/suspended/trial_expired
    subscription_plan: str | None = None


class QuotaUpdate(BaseModel):
    monthly_chat_cnt: int | None = None
    monthly_tokens: int | None = None
    monthly_embed_cnt: int | None = None


@router.get("/tenants")
def list_tenants(
    db: Session = Depends(get_db),
    user: dict = Depends(require_super_admin),
):
    """List all tenants with quota info."""
    companies = (
        db.query(Company)
        .filter(Company.deleted_at == None)
        .order_by(Company.company_id)
        .all()
    )

    result = []
    for c in companies:
        quota = db.query(TenantQuota).filter(TenantQuota.company_id == c.company_id).first()
        result.append({
            "company_id": c.company_id,
            "company_name": c.company_name,
            "status": getattr(c, "status", "active"),
            "subscription_plan": c.subscription_plan,
            "is_active": c.is_active,
            "quota": {
                "monthly_chat_cnt": quota.monthly_chat_cnt if quota else None,
                "monthly_tokens": quota.monthly_tokens if quota else None,
                "monthly_embed_cnt": quota.monthly_embed_cnt if quota else None,
            } if quota else None,
        })

    return {"tenants": result, "total": len(result)}


@router.put("/tenants/{company_id}")
def update_tenant(
    company_id: int,
    data: TenantUpdate,
    db: Session = Depends(get_db),
    user: dict = Depends(require_super_admin),
):
    """Update tenant status or subscription plan."""
    company = db.query(Company).filter(Company.company_id == company_id).first()
    if not company:
        raise HTTPException(status_code=404, detail="회사를 찾을 수 없습니다.")

    if data.status is not None:
        company.status = data.status
    if data.subscription_plan is not None:
        company.subscription_plan = data.subscription_plan

    db.commit()
    return {"success": True, "message": "테넌트 정보가 업데이트되었습니다."}


@router.put("/tenants/{company_id}/quota")
def update_quota(
    company_id: int,
    data: QuotaUpdate,
    db: Session = Depends(get_db),
    user: dict = Depends(require_super_admin),
):
    """Update or create quota for a tenant."""
    quota = db.query(TenantQuota).filter(TenantQuota.company_id == company_id).first()
    if not quota:
        quota = TenantQuota(company_id=company_id)
        db.add(quota)

    if data.monthly_chat_cnt is not None:
        quota.monthly_chat_cnt = data.monthly_chat_cnt
    if data.monthly_tokens is not None:
        quota.monthly_tokens = data.monthly_tokens
    if data.monthly_embed_cnt is not None:
        quota.monthly_embed_cnt = data.monthly_embed_cnt

    db.commit()
    return {"success": True, "message": "쿼터가 업데이트되었습니다."}


@router.post("/embeddings/rebuild")
def rebuild_embeddings(
    company_id: int | None = Query(None),
    db: Session = Depends(get_db),
    user: dict = Depends(require_super_admin),
):
    """Rebuild all embeddings (optionally for a specific company)."""
    stats = bulk_rebuild_embeddings(db, company_id)
    return {"success": True, "message": "임베딩 재생성 완료", **stats}


VALID_CATEGORIES = {"세금", "급여", "비용처리", "회계처리", "기타"}


@router.get("/qa/upload-template")
def download_upload_template(user: dict = Depends(require_super_admin)):
    """Download empty Excel template for QA bulk upload."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Q&A 업로드 양식"

    headers = [
        ("카테고리", 15),
        ("질문", 50),
        ("답변", 70),
        ("키워드", 20),
        ("별칭", 20),
        ("태그", 15),
        ("활성상태", 10),
    ]

    header_font = Font(name="맑은 고딕", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    for col_idx, (header_name, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    # Example row
    example = ["세금", "부가가치세 신고 기한은 언제인가요?", "부가가치세 신고 기한은 매 분기 종료 후 25일 이내입니다.", "부가세,신고,기한", "", "", "활성"]
    example_font = Font(name="맑은 고딕", size=10, color="999999")
    for col_idx, val in enumerate(example, 1):
        cell = ws.cell(row=2, column=col_idx, value=val)
        cell.font = example_font
        cell.border = thin_border

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=QA_upload_template.xlsx"},
    )


@router.post("/qa/upload")
def upload_qa_excel(
    company_id: int = Query(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: dict = Depends(require_super_admin),
):
    """Bulk-upload QA items from an Excel file."""
    if not file.filename or not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="xlsx 파일만 업로드 가능합니다.")

    # Verify company exists
    company = db.query(Company).filter(Company.company_id == company_id).first()
    if not company:
        raise HTTPException(status_code=404, detail="회사를 찾을 수 없습니다.")

    from openpyxl import load_workbook

    try:
        contents = file.file.read()
        wb = load_workbook(filename=io.BytesIO(contents), read_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="엑셀 파일을 읽을 수 없습니다.")

    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    created = 0
    skipped = 0
    failed = 0
    errors = []
    user_id = user.get("user_id")

    for idx, row in enumerate(rows, start=2):
        # Skip empty rows
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            skipped += 1
            continue

        # Extract columns (pad with None if row is shorter)
        padded = list(row) + [None] * (7 - len(row)) if len(row) < 7 else list(row)
        category = str(padded[0] or "").strip()
        question = str(padded[1] or "").strip()
        answer = str(padded[2] or "").strip()
        keywords = str(padded[3] or "").strip()
        aliases = str(padded[4] or "").strip()
        tags = str(padded[5] or "").strip()
        active_str = str(padded[6] or "활성").strip()

        # Validation
        row_errors = []
        if category not in VALID_CATEGORIES:
            row_errors.append(f"카테고리가 유효하지 않습니다 ('{category}')")
        if len(question) < 5:
            row_errors.append("질문이 5자 미만입니다")
        if len(answer) < 10:
            row_errors.append("답변이 10자 미만입니다")

        if row_errors:
            failed += 1
            errors.append(f"{idx}행: {', '.join(row_errors)}")
            continue

        is_active = active_str != "비활성"

        qa = QaKnowledge(
            company_id=company_id,
            category=category,
            question=question,
            answer=answer,
            keywords=keywords,
            aliases=aliases,
            tags=tags,
            is_active=is_active,
            created_by=user_id,
            updated_by=user_id,
        )
        db.add(qa)
        db.flush()  # Get qa_id for embedding

        try:
            upsert_qa_embedding(db, qa)
        except Exception as e:
            logger.warning("Embedding failed for row %d: %s", idx, e)

        created += 1

    db.commit()
    wb.close()

    return {
        "success": True,
        "total_rows": len(rows),
        "created": created,
        "skipped": skipped,
        "failed": failed,
        "errors": errors,
    }
