"""회사번호 1번 Q&A 데이터 전체를 엑셀로 내보내기"""

import sys
import os

sys.path.insert(0, os.path.dirname(__file__))

from sqlalchemy import create_engine, text
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from app.config import DATABASE_URL
from datetime import datetime


def export_qa_to_excel():
    # DB 연결
    connect_args = {}
    if DATABASE_URL.startswith("sqlite"):
        connect_args = {"check_same_thread": False}

    engine = create_engine(DATABASE_URL, connect_args=connect_args)

    # 데이터 조회
    query = text("""
        SELECT qa_id, company_id, category, question, answer,
               keywords, aliases, tags, is_active,
               view_count, used_count, created_at, updated_at
        FROM qa_knowledge
        WHERE company_id = 1
        ORDER BY qa_id ASC
    """)

    with engine.connect() as conn:
        rows = conn.execute(query).fetchall()

    print(f"조회된 Q&A 데이터: {len(rows)}건")

    if not rows:
        print("데이터가 없습니다.")
        return

    # 엑셀 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "Q&A 데이터 (회사 1번)"

    # 헤더 정의
    headers = [
        ("No", 6),
        ("QA ID", 8),
        ("카테고리", 15),
        ("질문", 50),
        ("답변", 70),
        ("키워드", 20),
        ("별칭", 20),
        ("태그", 15),
        ("활성상태", 10),
        ("조회수", 10),
        ("사용수", 10),
        ("생성일시", 20),
        ("수정일시", 20),
    ]

    # 스타일
    header_font = Font(name="맑은 고딕", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_font = Font(name="맑은 고딕", size=10)
    cell_align = Alignment(vertical="top", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="top")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    # 헤더 작성
    for col_idx, (header_name, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 데이터 작성
    for row_idx, row in enumerate(rows, 2):
        no = row_idx - 1
        qa_id, company_id, category, question, answer, keywords, aliases, tags, is_active, view_count, used_count, created_at, updated_at = row

        # 날짜 포맷
        def fmt_date(dt):
            if dt is None:
                return ""
            if isinstance(dt, str):
                return dt[:19]
            return dt.strftime("%Y-%m-%d %H:%M:%S")

        values = [
            no,
            qa_id,
            category,
            question,
            answer,
            keywords,
            aliases,
            tags,
            "활성" if is_active else "비활성",
            view_count or 0,
            used_count or 0,
            fmt_date(created_at),
            fmt_date(updated_at),
        ]

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = cell_font
            cell.border = thin_border
            if col_idx in (1, 2, 9, 10, 11):  # No, QA ID, 활성상태, 조회수, 사용수
                cell.alignment = center_align
            else:
                cell.alignment = cell_align

        # 짝수 행 배경색
        if row_idx % 2 == 0:
            even_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = even_fill

    # 행 높이 고정
    ws.row_dimensions[1].height = 30
    for r in range(2, len(rows) + 2):
        ws.row_dimensions[r].height = 45

    # 필터 설정
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"

    # 틀 고정 (헤더 고정)
    ws.freeze_panes = "A2"

    # 파일 저장
    output_dir = os.path.join(os.path.dirname(__file__), "data")
    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"QA_데이터_회사1번_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)

    wb.save(filepath)
    print(f"엑셀 파일 생성 완료: {filepath}")
    print(f"총 {len(rows)}건의 Q&A 데이터가 저장되었습니다.")


if __name__ == "__main__":
    export_qa_to_excel()
